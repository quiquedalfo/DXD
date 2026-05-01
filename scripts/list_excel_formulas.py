"""
Lista todas las fórmulas de un libro Excel (.xlsx o .xlsm).
Uso: python scripts/list_excel_formulas.py [ruta_al_archivo]
"""
import sys
from pathlib import Path

# Añadir raíz del proyecto al path
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

try:
    import openpyxl
except ImportError:
    print("Instala openpyxl: pip install openpyxl")
    sys.exit(1)


def list_formulas(filepath: Path) -> None:
    # load_workbook con data_only=False para ver fórmulas (no resultados)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=False, keep_vba=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        formulas_in_sheet = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas_in_sheet.append((cell.coordinate, cell.value))
        if formulas_in_sheet:
            print(f"\n--- Hoja: {sheet_name} ---")
            for coord, formula in formulas_in_sheet:
                print(f"  {coord}: {formula}")
    wb.close()


def main() -> None:
    if len(sys.argv) >= 2:
        filepath = Path(sys.argv[1])
    else:
        filepath = ROOT / "DXD.xlsm"
    if not filepath.exists():
        print(f"No existe: {filepath}")
        sys.exit(1)
    print(f"Fórmulas en: {filepath}")
    list_formulas(filepath)


if __name__ == "__main__":
    main()
